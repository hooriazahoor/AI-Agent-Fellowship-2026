"""
Generates docs/agent_evaluation.xlsx from tests/evaluation/results.json.
Run after run_evaluation.py.
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
results_path = ROOT / "tests" / "evaluation" / "results.json"
out_path = ROOT / "docs" / "agent_evaluation.xlsx"

data = json.loads(results_path.read_text())
results = data["results"]
recovery = data["recovery"]
metrics = data["metrics"]

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1E2530")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
PASS_FILL = PatternFill("solid", fgColor="E4F2EE")
FAIL_FILL = PatternFill("solid", fgColor="FBEAE8")
WRAP = Alignment(wrap_text=True, vertical="top")

wb = Workbook()

# ---------------------------------------------------------------------------
# Sheet 1: Test Cases
# ---------------------------------------------------------------------------
ws = wb.active
ws.title = "Test Cases"
headers = ["ID", "Category", "User Request", "Expected Tool", "Expected Arguments",
           "Approval Required", "Expected Outcome", "Actual Outcome", "Pass/Fail", "Notes"]
ws.append(headers)
for col in range(1, len(headers) + 1):
    c = ws.cell(row=1, column=col)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = WRAP

widths = [6, 14, 34, 26, 22, 12, 34, 34, 10, 34]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

for r in results:
    expected_tool = r["expected_tool"]
    expected_tool_str = ", ".join(expected_tool) if isinstance(expected_tool, list) else (expected_tool or "(none)")
    row = [
        r["id"], r["category"], r["user_request"], expected_tool_str,
        json.dumps(r["expected_args"]) if r["expected_args"] else "(none)",
        "Yes" if r["approval_required"] else "No",
        r["expected_outcome"], r["actual_outcome"],
        "PASS" if r["pass"] else "FAIL", r["notes"],
    ]
    ws.append(row)
    row_idx = ws.max_row
    fill = PASS_FILL if r["pass"] else FAIL_FILL
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = Font(name=FONT, size=9)
        cell.alignment = WRAP
        cell.fill = fill
ws.freeze_panes = "A2"

# ---------------------------------------------------------------------------
# Sheet 2: Metrics Summary
# ---------------------------------------------------------------------------
ws2 = wb.create_sheet("Metrics Summary")
ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 12
ws2.column_dimensions["E"].width = 40

headers2 = ["Metric", "Target", "Actual", "Met?", "Definition"]
ws2.append(headers2)
for col in range(1, 6):
    c = ws2.cell(row=1, column=col)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = WRAP

metric_rows = [
    ("Tool Selection Accuracy", "≥ 85%", f"{metrics['tool_selection_accuracy_pct']}%",
     "% of cases where the correct tool (or correct 'no tool' for direct-answer cases) was selected."),
    ("Argument Accuracy", "≥ 80%", f"{metrics['argument_accuracy_pct']}%",
     "% of tool calls (among cases with specific expected arguments) whose arguments matched."),
    ("Task Completion Rate", "≥ 80%", f"{metrics['task_completion_rate_pct']}%",
     "% of requests that reached the expected terminal state (tool+args+approval all correct, no crash)."),
    ("Approval Compliance", "= 100%", f"{metrics['approval_compliance_pct']}%",
     "% of write/sensitive actions where approval was correctly requested before execution."),
    ("Invalid Action Rate", "< 10%", f"{metrics['invalid_action_rate_pct']}%",
     "% of runs where the agent called an unexpected/incorrect tool."),
    ("Average Response Time", "(informational)", f"{metrics['average_response_time_ms']} ms",
     "Mean wall-clock duration per agent run (offline heuristic mode, local SQLite)."),
    ("Recovery Rate", "(informational)", f"{metrics['recovery_rate_pct']}%",
     "% of recoverable tool failures (transient, within retry budget) that succeeded after retry."),
]

targets_numeric = {"Tool Selection Accuracy": 85, "Argument Accuracy": 80,
                    "Task Completion Rate": 80, "Approval Compliance": 100, "Invalid Action Rate": 10}

for name, target, actual, definition in metric_rows:
    met = ""
    if name in targets_numeric:
        actual_val = float(actual.replace("%", ""))
        if name == "Invalid Action Rate":
            met = "YES" if actual_val < targets_numeric[name] else "NO"
        else:
            met = "YES" if actual_val >= targets_numeric[name] else "NO"
    ws2.append([name, target, actual, met, definition])
    row_idx = ws2.max_row
    fill = PASS_FILL if met == "YES" else (FAIL_FILL if met == "NO" else None)
    for col in range(1, 6):
        cell = ws2.cell(row=row_idx, column=col)
        cell.font = Font(name=FONT, size=10)
        cell.alignment = WRAP
        if fill:
            cell.fill = fill

ws2.append([])
ws2.append(["Total cases", metrics["total_cases"]])
ws2.append(["Passed", metrics["passed_cases"]])
ws2.append(["Failed", metrics["failed_cases"]])
for row_idx in range(ws2.max_row - 2, ws2.max_row + 1):
    ws2.cell(row=row_idx, column=1).font = Font(name=FONT, bold=True, size=10)
    ws2.cell(row=row_idx, column=2).font = Font(name=FONT, size=10)

# ---------------------------------------------------------------------------
# Sheet 3: Recovery Rate Evidence
# ---------------------------------------------------------------------------
ws3 = wb.create_sheet("Recovery Test Evidence")
ws3.column_dimensions["A"].width = 34
ws3.column_dimensions["B"].width = 14
ws3.column_dimensions["C"].width = 14
ws3.append(["Scenario", "Recovered?", "Attempts Made"])
for col in range(1, 4):
    c = ws3.cell(row=1, column=col)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
for case in recovery["cases"]:
    ws3.append([case["scenario"], "YES" if case["recovered"] else "NO", case["attempts"]])
    row_idx = ws3.max_row
    fill = PASS_FILL if case["recovered"] else FAIL_FILL
    for col in range(1, 4):
        cell = ws3.cell(row=row_idx, column=col)
        cell.font = Font(name=FONT, size=10)
        cell.fill = fill
ws3.append([])
ws3.append(["Note: simulated via deliberate fault injection (a test-only tool that fails N times "
            "then succeeds), run through the real _execute_with_retries pipeline with "
            "MAX_TOOL_RETRIES=2. Both scenarios are within the recoverable retry budget."])

wb.save(out_path)
print(f"Saved {out_path}")